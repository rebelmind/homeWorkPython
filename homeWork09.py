import os
os.chdir("F:/Python/ClassWork/homeWorkPython")
print("Текущая рабочая директория:", os.getcwd())


from openpyxl import Workbook, load_workbook
data = []

for i in range(1, 4):  
    file_path = f"F:/Python/ClassWork/homeWorkPython/{i}.xlsx" 
    wb = load_workbook(file_path)
    print(f"Файл {file_path} успешно загружен!")
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        # print(row[0].value)
        data.append(row[0].value)


data.sort(reverse= True)

wb = Workbook()
ws = wb.active
ws.title = "Strted"
ws.append(["Data"])

for val in data:
    print(val)
    ws.append([val])

from openpyxl.styles import Font, Border, Side
font = Font(name="Arial", size=12, bold=True)
border = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.font = font
        cell.border = border

wb.save("newExcel.xlsx")